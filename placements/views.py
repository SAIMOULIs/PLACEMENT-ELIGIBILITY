import csv
import io
from django.http import HttpResponse
from django.db.models import Count, Avg, Q
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny

from .models import Student, Company, Shortlist, UploadBatch, SystemLog, User
from .serializers import (StudentSerializer, CompanySerializer,
                           ShortlistSerializer, UploadBatchSerializer, UserSerializer)
from .permissions import IsAdminOrOfficer, IsAdminOfficerOrReadOnly
from .services.eligibility_engine import run_eligibility_engine


# ── Students ──────────────────────────────────────────────────────
class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAdminOfficerOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        branch = self.request.query_params.get('branch')
        year   = self.request.query_params.get('year')
        cgpa   = self.request.query_params.get('min_cgpa')
        search = self.request.query_params.get('search')
        if branch:
            qs = qs.filter(branch=branch.upper())
        if year:
            qs = qs.filter(graduation_year=year)
        if cgpa:
            qs = qs.filter(cgpa__gte=cgpa)
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(roll_number__icontains=search))
        return qs

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrOfficer])
    def upload(self, request):
        """Bulk upload students via CSV or Excel file."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=400)

        file_name = file.name
        batch = UploadBatch.objects.create(
            file_name=file_name,
            uploaded_by=request.user,
            status='processing',
        )

        success, failed, errors = 0, 0, []

        try:
            if file_name.endswith('.csv'):
                decoded = file.read().decode('utf-8')
                reader  = csv.DictReader(io.StringIO(decoded))
                rows    = list(reader)
            elif file_name.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb   = openpyxl.load_workbook(file)
                ws   = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                batch.status = 'failed'
                batch.save()
                return Response({'error': 'Only CSV and Excel files are supported.'}, status=400)

            for i, row in enumerate(rows, start=2):
                try:
                    Student.objects.update_or_create(
                        roll_number=str(row.get('roll_number', '')).strip(),
                        defaults={
                            'name':            str(row.get('name', '')).strip(),
                            'email':           str(row.get('email', '')).strip(),
                            'cgpa':            float(row.get('cgpa', 0)),
                            'backlogs':        int(row.get('backlogs', 0)),
                            'branch':          str(row.get('branch', '')).strip().upper(),
                            'skills':          str(row.get('skills', '')).strip(),
                            'graduation_year': int(row.get('graduation_year', 2025)),
                            'phone':           str(row.get('phone', '')).strip(),
                            'upload_batch':    batch,
                        }
                    )
                    success += 1
                except Exception as e:
                    failed += 1
                    errors.append(f"Row {i}: {str(e)}")

        except Exception as e:
            batch.status = 'failed'
            batch.error_log = str(e)
            batch.save()
            return Response({'error': str(e)}, status=500)

        batch.total_records   = success + failed
        batch.success_records = success
        batch.failed_records  = failed
        batch.status          = 'completed'
        batch.error_log       = '\n'.join(errors)
        batch.save()

        SystemLog.objects.create(
            log_type='upload',
            message=f"Uploaded {file_name}: {success} success, {failed} failed.",
            user=request.user,
        )

        return Response({
            'batch_id':    batch.id,
            'file':        file_name,
            'total':       success + failed,
            'success':     success,
            'failed':      failed,
            'errors':      errors[:10],
        }, status=201)


# ── Companies ─────────────────────────────────────────────────────
class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [IsAdminOfficerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminOrOfficer])
    def generate_shortlist(self, request, pk=None):
        """Run eligibility engine and generate shortlist for this company."""
        company = self.get_object()
        result  = run_eligibility_engine(company, user=request.user)
        return Response(result, status=200)

    @action(detail=True, methods=['get'])
    def shortlist(self, request, pk=None):
        """Get the current shortlist for this company."""
        company    = self.get_object()
        shortlists = Shortlist.objects.filter(company=company).select_related('student')
        serializer = ShortlistSerializer(shortlists, many=True)
        return Response({'company': company.name, 'count': shortlists.count(),
                         'shortlist': serializer.data})


# ── Shortlists ────────────────────────────────────────────────────
class ShortlistViewSet(viewsets.ModelViewSet):
    queryset = Shortlist.objects.select_related('student', 'company').all()
    serializer_class = ShortlistSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs         = super().get_queryset()
        company_id = self.request.query_params.get('company')
        student_id = self.request.query_params.get('student')
        s_status   = self.request.query_params.get('status')
        if company_id:
            qs = qs.filter(company_id=company_id)
        if student_id:
            qs = qs.filter(student_id=student_id)
        if s_status:
            qs = qs.filter(status=s_status)
        return qs

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export shortlist as CSV file."""
        company_id = request.query_params.get('company')
        qs = self.get_queryset()
        if company_id:
            qs = qs.filter(company_id=company_id)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="shortlist.csv"'
        writer = csv.writer(response)
        writer.writerow(['Roll Number', 'Name', 'Email', 'CGPA',
                         'Backlogs', 'Branch', 'Skills', 'Company', 'Status', 'Date'])
        for item in qs:
            writer.writerow([
                item.student.roll_number, item.student.name,
                item.student.email,       item.student.cgpa,
                item.student.backlogs,    item.student.branch,
                item.student.skills,      item.company.name,
                item.status,              item.shortlisted_at.date(),
            ])
        return response


# ── Upload Batches ────────────────────────────────────────────────
class UploadBatchViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UploadBatch.objects.all()
    serializer_class = UploadBatchSerializer
    permission_classes = [IsAdminOrOfficer]


# ── Dashboard / Analytics ─────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    total_students  = Student.objects.count()
    total_companies = Company.objects.count()
    total_shortlists = Shortlist.objects.count()
    avg_cgpa        = Student.objects.aggregate(avg=Avg('cgpa'))['avg'] or 0

    branch_dist = list(
        Student.objects.values('branch').annotate(count=Count('id')).order_by('-count')
    )
    company_stats = list(
        Company.objects.annotate(shortlisted=Count('shortlists')).values(
            'name', 'package', 'shortlisted'
        ).order_by('-shortlisted')[:10]
    )

    return Response({
        'total_students':   total_students,
        'total_companies':  total_companies,
        'total_shortlists': total_shortlists,
        'avg_cgpa':         round(float(avg_cgpa), 2),
        'branch_distribution': branch_dist,
        'top_companies':    company_stats,
    })


# ── Register User (Admin only) ────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def register_user(request):
    if request.user.role != 'admin':
        return Response({'error': 'Only admins can create users.'}, status=403)
    data = request.data
    try:
        user = User.objects.create_user(
            username=data['username'],
            password=data['password'],
            email=data.get('email', ''),
            role=data.get('role', 'student'),
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
        )
        return Response(UserSerializer(user).data, status=201)
    except Exception as e:
        return Response({'error': str(e)}, status=400)
